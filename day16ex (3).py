# -*- coding: utf-8 -*-
"""
Created on Tue Jun 25 19:23:36 2024

@author: USER
"""

import openpyxl;

wb = openpyxl.Workbook();
ws = wb.active;

row = ["國","英","數","合計"];
ws.append(row);

row1 = [100,80,66];
ws.append(row1);

ws['C4'] = '=sum(A2:C2)';
ws['C5'] = '=average(A2:C2)';
ws['C6'] = '=max(A2:C2)';
ws['C7'] = '=min(A2:C2)';

wb.save('demo4.xlsx');
