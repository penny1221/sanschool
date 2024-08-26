# -*- coding: utf-8 -*-
"""
Created on Thu Jun 20 21:03:59 2024

@author: USER
"""

import openpyxl;

file = "sales.xlsx";

wb = openpyxl.load_workbook(file,data_only=True);

ws = wb.active; #指定工作表 ex:wb['工作表名稱']

# print("目前工作表",ws.title);
# print(ws['A4'].value);
# print(ws['C2'].value);

for i in range(3,ws.max_row+1):
    for j in range(1,ws.max_column+1):
        print(ws.cell(i,j).value,end=",");
    print();